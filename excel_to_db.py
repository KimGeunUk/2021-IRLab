import pymysql                      # MySQL USE
from openpyxl import Workbook       # EXCEL USE
from openpyxl import load_workbook

def update_test():
    conn = pymysql.connect(host='localhost', user='root', password='irlab', db='irlab', charset='utf8')

    try:
        with conn.cursor() as cursor:
            # EXCEL DB LOAD
            wb = load_workbook('MODI2.xlsx', data_only=True)
            ws = wb['MODI1']  # worksheet 지정

            iter_rows = iter(ws.rows)
            next(iter_rows)

            sql = 'update new_foods set product_num=%s where name=%s'

            # EXCEL DB를 한 행씩 읽어 필요한 DATA만 INSERT 실행
            for row in iter_rows:
                #print(row[1].value, row[2].value)
                cursor.execute(sql, (row[1].value, row[2].value))

        conn.commit()
    finally:
        conn.close()

def insert_excel_to_db():
    # MYSQL CONNET
    conn = pymysql.connect(host='localhost', user='root', password='irlab', db='irlab', charset='utf8')

    try:
        with conn.cursor() as cursor:
            # EXCEL DB LOAD 원본 INSERT
            wb = load_workbook('./DB/FOOD_DB_ALL_MODI.xlsx', data_only=True)
            ws = wb['정제']   # EXCEL worksheet 지정

            # 불러온 EXCEL 데이터를 한 행씩 읽어 iter_rows 변수에 저장
            iter_rows = iter(ws.rows)
            # 데이터는 5번째 행부터 시작하여, 4행을 넘긴다.
            for i in range(1):
                next(iter_rows)

            # INSERT SQL문 작성
            sql = 'insert into foods values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'

            # EXCEL DB를 한 행씩 읽어 필요한 DATA만 INSERT 실행
            # enumerate 함수는 몇번째 반복문인지 확인 하기 위해 인덱스 번호(idx)를 같이 반환해 준다.
            for idx, row in enumerate(iter_rows):
                # 시간이 너무 오래 걸리고, RAM 부족으로 실행이 안된다.
                # # EXCEL 파일의 6번째 열(row[6])은 제작년도를 의미한다. 이 값이 2016-2019 년도 사이면 그 행을 삭제한다.
                # if int(row[6].value) >= 2016 and int(row[6].value) <= 2019:
                #     ws.delete_rows(idx+1)
                # elif row[3].value != '가공식품':
                #     ws.delete_rows(idx + 1)
                #
                # # EXCEL 파일의 11번째 열부터 140번째 열은 데이터를 의미한다. 이 값안에 문자열이 있으면 문자열을 제거한다.
                # for i in range(11, 140):
                #     if '미만' in str(row[i].value):
                #         row[i].value = row[i].value.replace('미만', '')
                #     elif ' 미만' in str(row[i].value):
                #         row[i].value = row[i].value.replace(' 미만', '')
                #     elif 'g미만' in str(row[i].value):
                #         row[i].value = row[i].value.replace('g미만', '')
                #     elif 'g 미만' in str(row[i].value):
                #         row[i].value = row[i].value.replace('g 미만', '')

                # 작성해둔 SQL문의 형식에 맞게 데이터를 넣은 후 실행한다.
                cursor.execute(sql, (None, None, row[5].value, row[7].value, None, None, None, row[9].value, row[11].value,
                                     row[15].value, row[19].value, row[20].value, row[22].value, row[23].value,
                                     row[39].value, row[45].value, row[46].value, row[135].value, row[138].value))

                # cursor.execute(sql, (None, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value,
                #                      row[7].value, row[8].value, row[9].value, row[10].value, row[11].value, row[12].value,
                #                      row[13].value, row[14].value, row[15].value, row[16].value, row[17].value, row[18].value))

            # COMMIT
            conn.commit()
    finally:
        conn.close()
        wb.close()

if __name__ == '__main__':
    insert_excel_to_db()
